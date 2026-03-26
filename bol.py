from openpyxl import load_workbook
from pathlib import Path


INPUT_FILE = "girdi.xlsx"
OUTPUT_FILE = "cikti.xlsx"
SHEET_NAME = None          # boş olursa aktif sayfa kullanılır
SOURCE_CELL = "A1"         # Bölünecek hücre
TARGET_COLUMN = 1          # 1 = A sütunu


def split_cell_lines_to_rows(
    input_file: str,
    output_file: str,
    source_cell: str = "A1",
    target_column: int = 1,
    sheet_name: str | None = None,
) -> None:
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")

    wb = load_workbook(input_file)
    ws = wb[sheet_name] if sheet_name else wb.active

    cell_value = ws[source_cell].value

    if cell_value is None:
        raise ValueError(f"{source_cell} hücresi boş.")

    lines = [line.strip() for line in str(cell_value).splitlines() if line.strip()]

    if not lines:
        raise ValueError(f"{source_cell} hücresinde ayrıştırılacak satır bulunamadı.")

    for row_index, line in enumerate(lines, start=1):
        ws.cell(row=row_index, column=target_column, value=line)

    wb.save(output_file)
    print(f"Tamamlandı: {output_file}")


if __name__ == "__main__":
    split_cell_lines_to_rows(
        input_file=INPUT_FILE,
        output_file=OUTPUT_FILE,
        source_cell=SOURCE_CELL,
        target_column=TARGET_COLUMN,
        sheet_name=SHEET_NAME,
    )
